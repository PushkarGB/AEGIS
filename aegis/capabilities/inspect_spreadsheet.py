"""Deterministic spreadsheet inspection capability using openpyxl."""

from __future__ import annotations

import datetime
from decimal import Decimal
import math
import os
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException
from pydantic import BaseModel, ConfigDict, Field

from aegis.capabilities.base import (
    Capability,
    CapabilityContract,
    CapabilityKind,
    CapabilityMetadata,
)
from aegis.schemas import (
    CapabilityRequest,
    CapabilityResult,
    CapabilityResultStatus,
    JsonObject,
    Observation,
)


class ColumnInfo(BaseModel):
    """Structured inspection details for a single spreadsheet column."""

    model_config = ConfigDict(extra="forbid", frozen=True)

    name: str = Field(min_length=1)
    letter: str = Field(min_length=1)
    index: int = Field(ge=1)
    inferred_type: str = Field(min_length=1)
    is_numeric: bool
    sample_values: list[Any] = Field(default_factory=list)
    non_empty_count: int = Field(ge=0)
    null_count: int = Field(ge=0)
    min_value: float | int | None = None
    max_value: float | int | None = None


class SheetInfo(BaseModel):
    """Structured inspection details for a single worksheet."""

    model_config = ConfigDict(extra="forbid", frozen=True)

    name: str = Field(min_length=1)
    index: int = Field(ge=1)
    is_active: bool
    is_empty: bool
    total_rows: int = Field(ge=0)
    data_row_count: int = Field(ge=0)
    column_count: int = Field(ge=0)
    columns: list[str] = Field(default_factory=list)
    column_details: list[ColumnInfo] = Field(default_factory=list)
    numeric_fields: list[str] = Field(default_factory=list)
    representative_values: dict[str, list[Any]] = Field(default_factory=dict)
    preview_rows: list[dict[str, Any]] = Field(default_factory=list)


class WorkbookMetadata(BaseModel):
    """Basic file and workbook metadata."""

    model_config = ConfigDict(extra="forbid", frozen=True)

    file_name: str = Field(min_length=1)
    file_path: str = Field(min_length=1)
    file_size_bytes: int = Field(ge=0)
    sheet_names: list[str] = Field(default_factory=list)
    sheet_count: int = Field(ge=0)
    active_sheet: str | None = None
    title: str | None = None
    creator: str | None = None
    created: str | None = None
    modified: str | None = None


class WorkbookInspection(BaseModel):
    """Complete structured inspection result for an Excel workbook."""

    model_config = ConfigDict(extra="forbid", frozen=True)

    metadata: WorkbookMetadata
    sheets: list[SheetInfo] = Field(default_factory=list)
    sheet_names: list[str] = Field(default_factory=list)
    total_sheets: int = Field(ge=0)
    columns: list[str] = Field(default_factory=list)
    row_count: int = Field(ge=0)
    numeric_fields: list[str] = Field(default_factory=list)
    representative_values: dict[str, list[Any]] = Field(default_factory=dict)
    preview_rows: list[dict[str, Any]] = Field(default_factory=list)

    def to_dict(self) -> JsonObject:
        """Convert inspection model to a JSON-compatible dictionary."""
        return self.model_dump(mode="json")


def _json_safe_value(value: Any) -> Any:
    """Normalize cell values into JSON-serializable primitives."""
    if value is None:
        return None
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return int(value) if value % 1 == 0 else float(value)
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return str(value)
        return value
    if isinstance(value, (int, bool, str)):
        return value
    return str(value)


def _infer_column_type(non_null_values: list[Any]) -> tuple[str, bool]:
    """Deterministically infer the column data type and numeric status."""
    if not non_null_values:
        return "empty", False

    # Check boolean first since bool is a subclass of int in Python
    if all(isinstance(v, bool) for v in non_null_values):
        return "boolean", False

    # Check pure integer
    if all(isinstance(v, int) and not isinstance(v, bool) for v in non_null_values):
        return "integer", True

    # Check numeric (float or int)
    if all(
        isinstance(v, (int, float, Decimal)) and not isinstance(v, bool)
        for v in non_null_values
    ):
        return "float", True

    # Check date / datetime
    if all(isinstance(v, (datetime.datetime, datetime.date)) for v in non_null_values):
        return "datetime", False

    return "string", False


def inspect_spreadsheet(
    file_path: str | Path,
    max_sample_values: int = 5,
    max_preview_rows: int = 5,
) -> WorkbookInspection:
    """Deterministically inspect an Excel workbook using openpyxl.

    Extracts sheet schemas, column definitions, data row counts,
    representative values, numeric field identification, and workbook metadata.
    """
    path = Path(file_path).resolve()
    if not path.exists():
        raise FileNotFoundError(f"Spreadsheet file not found: {path}")
    if not path.is_file():
        raise ValueError(f"Path is not a regular file: {path}")

    try:
        wb = openpyxl.load_workbook(filename=str(path), data_only=True)
    except Exception as exc:
        raise ValueError(f"Failed to open workbook '{path.name}': {exc}") from exc

    try:
        sheet_names = list(wb.sheetnames)
        active_sheet_title = wb.active.title if wb.active is not None else (
            sheet_names[0] if sheet_names else None
        )

        # Basic metadata
        wb_props = wb.properties
        created_str = (
            wb_props.created.isoformat()
            if wb_props and wb_props.created
            else None
        )
        modified_str = (
            wb_props.modified.isoformat()
            if wb_props and wb_props.modified
            else None
        )

        metadata = WorkbookMetadata(
            file_name=path.name,
            file_path=str(path),
            file_size_bytes=path.stat().st_size,
            sheet_names=sheet_names,
            sheet_count=len(sheet_names),
            active_sheet=active_sheet_title,
            title=wb_props.title if wb_props and wb_props.title else None,
            creator=wb_props.creator if wb_props and wb_props.creator else None,
            created=created_str,
            modified=modified_str,
        )

        sheets_info: list[SheetInfo] = []

        for sheet_idx, sheet in enumerate(wb.worksheets, start=1):
            sheet_name = sheet.title
            is_active = (sheet_name == active_sheet_title)

            # Read all rows as tuples
            raw_rows = list(sheet.iter_rows(values_only=True))

            # Filter out trailing completely empty rows
            non_empty_row_tuples = [
                row for row in raw_rows
                if any(cell is not None and str(cell).strip() != "" for cell in row)
            ]

            if not non_empty_row_tuples:
                sheets_info.append(
                    SheetInfo(
                        name=sheet_name,
                        index=sheet_idx,
                        is_active=is_active,
                        is_empty=True,
                        total_rows=0,
                        data_row_count=0,
                        column_count=0,
                        columns=[],
                        column_details=[],
                        numeric_fields=[],
                        representative_values={},
                        preview_rows=[],
                    )
                )
                continue

            header_raw = non_empty_row_tuples[0]
            data_raw_rows = non_empty_row_tuples[1:]

            # Determine column count from the widest row
            max_cols = max(len(row) for row in non_empty_row_tuples)

            # Build column headers with deduplication and fallback
            headers: list[str] = []
            seen_headers: dict[str, int] = {}
            for col_i in range(max_cols):
                raw_header = header_raw[col_i] if col_i < len(header_raw) else None
                if raw_header is not None and str(raw_header).strip() != "":
                    header_name = str(raw_header).strip()
                else:
                    col_letter = get_column_letter(col_i + 1)
                    header_name = f"column_{col_letter}"

                if header_name in seen_headers:
                    seen_headers[header_name] += 1
                    header_name = f"{header_name}_{seen_headers[header_name]}"
                else:
                    seen_headers[header_name] = 0

                headers.append(header_name)

            column_details: list[ColumnInfo] = []
            numeric_fields: list[str] = []
            representative_values: dict[str, list[Any]] = {}

            for col_i, header_name in enumerate(headers):
                col_letter = get_column_letter(col_i + 1)
                col_values = [
                    row[col_i] if col_i < len(row) else None
                    for row in data_raw_rows
                ]

                non_null_raw = [
                    v for v in col_values
                    if v is not None and (not isinstance(v, str) or v.strip() != "")
                ]
                null_count = len(col_values) - len(non_null_raw)
                non_empty_count = len(non_null_raw)

                inferred_type, is_numeric = _infer_column_type(non_null_raw)
                if is_numeric:
                    numeric_fields.append(header_name)

                # Compute min/max for numeric columns
                min_val: float | int | None = None
                max_val: float | int | None = None
                if is_numeric and non_null_raw:
                    num_vals = [
                        float(v) if isinstance(v, (int, float, Decimal)) else v
                        for v in non_null_raw
                    ]
                    min_num = min(num_vals)
                    max_num = max(num_vals)
                    min_val = int(min_num) if min_num % 1 == 0 else min_num
                    max_val = int(max_num) if max_num % 1 == 0 else max_num

                # Sample representative non-null values
                sample_vals = [
                    _json_safe_value(v)
                    for v in non_null_raw[:max_sample_values]
                ]

                col_info = ColumnInfo(
                    name=header_name,
                    letter=col_letter,
                    index=col_i + 1,
                    inferred_type=inferred_type,
                    is_numeric=is_numeric,
                    sample_values=sample_vals,
                    non_empty_count=non_empty_count,
                    null_count=null_count,
                    min_value=min_val,
                    max_value=max_val,
                )
                column_details.append(col_info)
                representative_values[header_name] = sample_vals

            # Preview rows
            preview_rows: list[dict[str, Any]] = []
            for row in data_raw_rows[:max_preview_rows]:
                row_dict: dict[str, Any] = {}
                for col_i, header_name in enumerate(headers):
                    val = row[col_i] if col_i < len(row) else None
                    row_dict[header_name] = _json_safe_value(val)
                preview_rows.append(row_dict)

            sheets_info.append(
                SheetInfo(
                    name=sheet_name,
                    index=sheet_idx,
                    is_active=is_active,
                    is_empty=False,
                    total_rows=len(non_empty_row_tuples),
                    data_row_count=len(data_raw_rows),
                    column_count=len(headers),
                    columns=headers,
                    column_details=column_details,
                    numeric_fields=numeric_fields,
                    representative_values=representative_values,
                    preview_rows=preview_rows,
                )
            )

        # Determine primary sheet summary (active sheet or first non-empty sheet or first sheet)
        primary_sheet = next(
            (s for s in sheets_info if s.name == active_sheet_title and not s.is_empty),
            next(
                (s for s in sheets_info if not s.is_empty),
                sheets_info[0] if sheets_info else None,
            ),
        )

        return WorkbookInspection(
            metadata=metadata,
            sheets=sheets_info,
            sheet_names=sheet_names,
            total_sheets=len(sheet_names),
            columns=primary_sheet.columns if primary_sheet else [],
            row_count=primary_sheet.data_row_count if primary_sheet else 0,
            numeric_fields=primary_sheet.numeric_fields if primary_sheet else [],
            representative_values=primary_sheet.representative_values if primary_sheet else {},
            preview_rows=primary_sheet.preview_rows if primary_sheet else [],
        )
    finally:
        wb.close()


class InspectSpreadsheetCapability(Capability):
    """Deterministic capability for inspecting Excel workbooks with openpyxl."""

    def __init__(self) -> None:
        self._metadata = CapabilityMetadata(
            name="inspect_spreadsheet",
            kind=CapabilityKind.TOOL,
            description="Deterministically inspect an Excel spreadsheet and extract structured schema and metadata.",
            input_contract=CapabilityContract(
                json_schema={
                    "type": "object",
                    "properties": {
                        "workbook": {"type": "string", "description": "Path to the workbook file."},
                        "file_path": {"type": "string", "description": "Alternative path to the workbook file."},
                        "path": {"type": "string", "description": "Alternative path to the workbook file."},
                        "max_sample_values": {"type": "integer", "description": "Max representative values per column."},
                        "max_preview_rows": {"type": "integer", "description": "Max preview data rows to return."},
                    },
                }
            ),
            output_contract=CapabilityContract(
                json_schema={
                    "type": "object",
                    "properties": {
                        "metadata": {"type": "object"},
                        "sheets": {"type": "array"},
                        "sheet_names": {"type": "array"},
                        "total_sheets": {"type": "integer"},
                        "columns": {"type": "array"},
                        "row_count": {"type": "integer"},
                        "numeric_fields": {"type": "array"},
                        "representative_values": {"type": "object"},
                        "preview_rows": {"type": "array"},
                    },
                    "required": ["metadata", "sheets", "sheet_names", "columns", "row_count"],
                }
            ),
            input_modalities=("spreadsheet",),
        )

    @property
    def metadata(self) -> CapabilityMetadata:
        return self._metadata

    def execute(self, request: CapabilityRequest) -> CapabilityResult:
        # Resolve target workbook file path from inputs
        inputs = request.inputs
        target_path: str | None = None
        for key in ("workbook", "file_path", "path", "filepath", "file", "attachment"):
            val = inputs.get(key)
            if isinstance(val, str) and val.strip():
                target_path = val.strip()
                break

        # Fallback: search for any string input containing a spreadsheet extension or path
        if target_path is None:
            for val in inputs.values():
                if isinstance(val, str) and (
                    val.endswith(".xlsx")
                    or val.endswith(".xlsm")
                    or val.endswith(".xltx")
                    or val.endswith(".xltm")
                ):
                    target_path = val
                    break

        if target_path is None:
            return CapabilityResult(
                request_id=request.request_id,
                status=CapabilityResultStatus.FAILED,
                error="No workbook file path provided in capability request inputs.",
            )

        max_sample_values = int(inputs.get("max_sample_values", 5))
        max_preview_rows = int(inputs.get("max_preview_rows", 5))

        try:
            inspection = inspect_spreadsheet(
                file_path=target_path,
                max_sample_values=max_sample_values,
                max_preview_rows=max_preview_rows,
            )
        except Exception as exc:
            return CapabilityResult(
                request_id=request.request_id,
                status=CapabilityResultStatus.FAILED,
                error=f"Failed to inspect spreadsheet: {exc}",
            )

        output_data = inspection.to_dict()

        # Build concise observation for Controller recording
        sheet_summary_parts = []
        for s in inspection.sheets:
            if s.is_empty:
                sheet_summary_parts.append(f"Sheet '{s.name}' (empty)")
            else:
                num_str = f", numeric: [{', '.join(s.numeric_fields)}]" if s.numeric_fields else ""
                sheet_summary_parts.append(
                    f"Sheet '{s.name}' ({s.data_row_count} rows, cols: [{', '.join(s.columns)}]{num_str})"
                )

        obs_summary = (
            f"Workbook '{inspection.metadata.file_name}' contains {inspection.total_sheets} sheet(s): "
            + "; ".join(sheet_summary_parts)
            + "."
        )

        observation = Observation(
            source=self.metadata.name,
            kind="spreadsheet_inspection",
            summary=obs_summary,
            data={
                "file_name": inspection.metadata.file_name,
                "sheet_names": inspection.sheet_names,
                "total_sheets": inspection.total_sheets,
                "columns": inspection.columns,
                "row_count": inspection.row_count,
                "numeric_fields": inspection.numeric_fields,
                "representative_values": inspection.representative_values,
            },
            request_id=request.request_id,
        )

        return CapabilityResult(
            request_id=request.request_id,
            status=CapabilityResultStatus.SUCCEEDED,
            output=output_data,
            observations=[observation],
        )
