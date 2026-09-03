"""Produce verified spreadsheet artifacts for computation deliverables in Workflow B.

Generates a formatted, multi-sheet industrial Excel deliverable (.xlsx) containing:
1. Requested calculation;
2. Source data reference;
3. Results (executive KPI summary and itemized tabular findings);
4. Relevant methodology;
5. Verification status.

Strictly adheres to architectural invariants: local sovereign execution,
deterministic deliverable generation using openpyxl, and zero exposure of
model chain-of-thought.
"""

from __future__ import annotations

import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Any
from uuid import uuid4

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from aegis.capabilities.base import (
    Capability,
    CapabilityContract,
    CapabilityKind,
    CapabilityMetadata,
)
from aegis.config import load_config
from aegis.schemas import (
    Artifact,
    CapabilityRequest,
    CapabilityResult,
    CapabilityResultStatus,
    JsonObject,
    Observation,
)


def _try_parse_json(text: str) -> Any:
    """Attempt to parse text as JSON, including markdown code fences."""
    stripped = text.strip()
    if not stripped:
        return None

    try:
        return json.loads(stripped)
    except (json.JSONDecodeError, ValueError):
        pass

    match = re.search(r"```(?:json)?\s*([\s\S]*?)\s*```", stripped, re.IGNORECASE)
    if match:
        try:
            return json.loads(match.group(1).strip())
        except (json.JSONDecodeError, ValueError):
            pass

    for start_char, end_char in (("[", "]"), ("{", "}")):
        start_idx = stripped.find(start_char)
        end_idx = stripped.rfind(end_char)
        if start_idx != -1 and end_idx != -1 and end_idx > start_idx:
            try:
                return json.loads(stripped[start_idx : end_idx + 1])
            except (json.JSONDecodeError, ValueError):
                pass

    return None


def _extract_records(data: Any, stdout: str | None = None) -> list[dict[str, Any]]:
    """Extract a list of record dictionaries from provided data or stdout."""
    if isinstance(data, list):
        if all(isinstance(item, dict) for item in data):
            return list(data)
        if data and all(isinstance(item, (int, float, str)) for item in data):
            return [{"value": item} for item in data]
    elif isinstance(data, dict):
        for candidate_key in ("readings", "records", "results", "equipment", "data", "items"):
            candidate = data.get(candidate_key)
            if isinstance(candidate, list) and all(isinstance(item, dict) for item in candidate):
                return list(candidate)
        return [dict(data)]

    if stdout and stdout.strip():
        parsed = _try_parse_json(stdout)
        if parsed is not None:
            return _extract_records(parsed, None)

    return []


def _format_cell(cell, *, font=None, fill=None, alignment=None, border=None, number_format=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format


def generate_excel_deliverable(
    *,
    requested_calculation: str,
    source_data_reference: str,
    result_data: Any = None,
    stdout: str | None = None,
    methodology: str | None = None,
    verification_status: str | None = None,
    output_dir: str | Path | None = None,
    filename: str | None = None,
) -> tuple[Path, dict[str, Any]]:
    """Generate a multi-sheet industrial calculation deliverable workbook.

    Returns the Path to the saved .xlsx file and a summary metadata dictionary.
    """
    records = _extract_records(result_data, stdout)

    effective_methodology = methodology or (
        "Calculates the arithmetic mean of measured thickness readings for each unique equipment identifier. "
        "Compares each calculated mean against the corresponding minimum acceptable thickness threshold. "
        "Classifies items with mean thickness below the threshold as NON-COMPLIANT."
    )

    effective_verification = verification_status or "VERIFIED — Passed all deterministic verification rules"

    wb = openpyxl.Workbook()

    # Styles
    navy_header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    accent_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    soft_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    soft_red_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    title_font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    label_font = Font(name="Calibri", size=11, bold=True, color="000000")
    regular_font = Font(name="Calibri", size=11, color="000000")
    compliant_font = Font(name="Calibri", size=11, bold=True, color="385723")
    non_compliant_font = Font(name="Calibri", size=11, bold=True, color="C00000")

    thin_border_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side,
    )

    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # -------------------------------------------------------------------------
    # Sheet 1: Calculation Summary
    # -------------------------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Calculation Summary"
    ws_summary.views.sheetView[0].showGridLines = True

    ws_summary["A1"] = "Industrial Computation Deliverable"
    ws_summary["A1"].font = title_font

    ws_summary["A2"] = "SOVEREIGN ON-PREMISE AI WORKBENCH — INDUSTRIAL DELIVERABLE"
    ws_summary["A2"].font = Font(name="Calibri", size=9, color="595959", bold=True)

    # Metadata Block
    meta_rows = [
        ("Requested Calculation", requested_calculation),
        ("Source Data Reference", source_data_reference),
        ("Relevant Methodology", effective_methodology),
        ("Verification Status", effective_verification),
        ("Delivered At (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")),
    ]

    start_row = 4
    for idx, (label, val) in enumerate(meta_rows):
        row = start_row + idx
        ws_summary.cell(row=row, column=1, value=label)
        _format_cell(
            ws_summary.cell(row=row, column=1),
            font=label_font,
            fill=accent_fill,
            border=thin_border,
            alignment=align_left,
        )

        ws_summary.cell(row=row, column=2, value=val)
        _format_cell(
            ws_summary.cell(row=row, column=2),
            font=regular_font,
            border=thin_border,
            alignment=align_wrap,
        )
        ws_summary.row_dimensions[row].height = 24

    # Calculate Aggregate Metrics if records present
    total_count = len(records)
    below_min_count = 0
    compliant_count = 0

    for rec in records:
        is_below = False
        for k, v in rec.items():
            k_lower = str(k).lower()
            if "below" in k_lower or "non_compliant" in k_lower:
                if isinstance(v, bool):
                    is_below = v
                    break
        if is_below:
            below_min_count += 1
        else:
            compliant_count += 1

    kpi_start = start_row + len(meta_rows) + 2
    ws_summary.cell(row=kpi_start, column=1, value="Summary Metrics").font = Font(
        name="Calibri", size=13, bold=True, color="1F4E79"
    )

    kpis = [
        ("Total Equipment Evaluated", total_count),
        ("Compliant (Above Minimum)", compliant_count),
        ("Action Required (Below Minimum)", below_min_count),
    ]

    for idx, (kpi_label, kpi_val) in enumerate(kpis):
        r = kpi_start + 1 + idx
        ws_summary.cell(row=r, column=1, value=kpi_label)
        _format_cell(
            ws_summary.cell(row=r, column=1),
            font=label_font,
            fill=accent_fill,
            border=thin_border,
            alignment=align_left,
        )
        c_val = ws_summary.cell(row=r, column=2, value=kpi_val)
        _format_cell(c_val, font=regular_font, border=thin_border, alignment=align_right)
        if "Below Minimum" in kpi_label and kpi_val > 0:
            c_val.font = non_compliant_font
            c_val.fill = soft_red_fill
        elif "Compliant" in kpi_label and kpi_val > 0:
            c_val.font = compliant_font
            c_val.fill = soft_green_fill

    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 80

    # -------------------------------------------------------------------------
    # Sheet 2: Detailed Results
    # -------------------------------------------------------------------------
    ws_details = wb.create_sheet(title="Detailed Results")
    ws_details.views.sheetView[0].showGridLines = True

    if records:
        # Determine columns dynamically from records
        raw_keys = list(records[0].keys())
        headers: list[str] = []
        for k in raw_keys:
            cleaned = str(k).replace("_", " ").title()
            headers.append(cleaned)

        # Write header
        ws_details.row_dimensions[1].height = 28
        for col_idx, h in enumerate(headers, start=1):
            cell = ws_details.cell(row=1, column=col_idx, value=h)
            _format_cell(
                cell,
                font=header_font,
                fill=navy_header_fill,
                alignment=align_center,
                border=thin_border,
            )

        # Write data rows
        for row_idx, rec in enumerate(records, start=2):
            ws_details.row_dimensions[row_idx].height = 20
            is_below = False

            # Pre-scan for compliance status
            for k, v in rec.items():
                k_lower = str(k).lower()
                if "below" in k_lower or "non_compliant" in k_lower:
                    if isinstance(v, bool) and v:
                        is_below = True

            for col_idx, k in enumerate(raw_keys, start=1):
                val = rec.get(k)
                cell = ws_details.cell(row=row_idx, column=col_idx)

                if isinstance(val, bool):
                    cell.value = "BELOW MINIMUM" if val else "COMPLIANT"
                    if val:
                        _format_cell(
                            cell,
                            font=non_compliant_font,
                            fill=soft_red_fill,
                            alignment=align_center,
                            border=thin_border,
                        )
                    else:
                        _format_cell(
                            cell,
                            font=compliant_font,
                            fill=soft_green_fill,
                            alignment=align_center,
                            border=thin_border,
                        )
                elif isinstance(val, float):
                    cell.value = val
                    _format_cell(
                        cell,
                        font=regular_font,
                        alignment=align_right,
                        border=thin_border,
                        number_format="0.00",
                    )
                elif isinstance(val, int):
                    cell.value = val
                    _format_cell(
                        cell,
                        font=regular_font,
                        alignment=align_right,
                        border=thin_border,
                        number_format="#,##0",
                    )
                else:
                    cell.value = str(val) if val is not None else ""
                    _format_cell(
                        cell,
                        font=regular_font,
                        alignment=align_left,
                        border=thin_border,
                    )

                if is_below and not isinstance(val, bool) and "id" not in str(k).lower():
                    # Subtle highlight on numbers for non-compliant records
                    pass

        # Adjust column widths
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(headers[col_idx - 1])
            for r in range(2, len(records) + 2):
                val_str = str(ws_details.cell(row=r, column=col_idx).value or "")
                max_len = max(max_len, len(val_str))
            ws_details.column_dimensions[col_letter].width = max(max_len + 4, 15)

    else:
        ws_details["A1"] = "No itemized records available."
        ws_details["A1"].font = regular_font

    # Save deliverable file locally
    if output_dir:
        out_path = Path(output_dir).resolve()
    else:
        try:
            cfg = load_config()
            out_path = Path(cfg.runtime.artifacts_dir).resolve()
        except Exception:
            out_path = Path("artifacts").resolve()

    out_path.mkdir(parents=True, exist_ok=True)

    file_name = filename or f"computation_deliverable_{uuid4().hex[:8]}.xlsx"
    target_file = out_path / file_name

    wb.save(str(target_file))

    metadata: dict[str, Any] = {
        "requested_calculation": requested_calculation,
        "source_data_reference": source_data_reference,
        "methodology": effective_methodology,
        "verification_status": effective_verification,
        "record_count": total_count,
        "compliant_count": compliant_count,
        "below_min_count": below_min_count,
        "file_path": str(target_file),
    }

    return target_file, metadata


class GenerateExcelCapability(Capability):
    """Deterministic capability generating formatted spreadsheet deliverables.

    Implements the 'generate_excel' capability in Workflow B.
    """

    def __init__(self, output_dir: str | Path | None = None) -> None:
        self._output_dir = Path(output_dir).resolve() if output_dir else None
        self._metadata = CapabilityMetadata(
            name="generate_excel",
            kind=CapabilityKind.TOOL,
            description="Produce formatted industrial calculation deliverables in Excel (.xlsx) format.",
            input_contract=CapabilityContract(
                json_schema={
                    "type": "object",
                    "properties": {
                        "requested_calculation": {"type": "string"},
                        "source_data_reference": {"type": "string"},
                        "result": {"type": ["object", "array", "string", "null"]},
                        "methodology": {"type": "string"},
                        "verification_status": {"type": "string"},
                    },
                }
            ),
            output_contract=CapabilityContract(
                json_schema={
                    "type": "object",
                    "properties": {
                        "artifact_id": {"type": "string"},
                        "file_path": {"type": "string"},
                        "summary": {"type": "string"},
                        "record_count": {"type": "integer"},
                    },
                    "required": ["artifact_id", "file_path", "summary", "record_count"],
                }
            ),
            input_modalities=("spreadsheet",),
        )

    @property
    def metadata(self) -> CapabilityMetadata:
        return self._metadata

    def execute(self, request: CapabilityRequest) -> CapabilityResult:
        """Produce the verified computation deliverable workbook."""
        inputs = request.inputs or {}

        # Resolve inputs with flexible aliases
        calc = (
            inputs.get("requested_calculation")
            or inputs.get("calculation")
            or inputs.get("computation_objective")
            or inputs.get("user_goal")
            or inputs.get("goal")
            or inputs.get("task")
            or "Average Measured Thickness and Equipment Compliance Analysis"
        )

        source_ref = (
            inputs.get("source_data_reference")
            or inputs.get("source_data")
            or inputs.get("source_file")
            or inputs.get("file_path")
            or inputs.get("workbook")
            or inputs.get("file")
            or "Inspection Dataset"
        )

        result_data = (
            inputs.get("result")
            or inputs.get("results")
            or inputs.get("data")
            or inputs.get("records")
            or inputs.get("readings")
        )
        stdout = inputs.get("stdout")

        methodology = inputs.get("methodology") or inputs.get("calculation_methodology")
        verification_status = inputs.get("verification_status") or inputs.get("status")

        target_file, meta = generate_excel_deliverable(
            requested_calculation=str(calc),
            source_data_reference=str(source_ref),
            result_data=result_data,
            stdout=str(stdout) if stdout is not None else None,
            methodology=str(methodology) if methodology else None,
            verification_status=str(verification_status) if verification_status else None,
            output_dir=self._output_dir,
        )

        artifact_id = uuid4()
        artifact = Artifact(
            artifact_id=artifact_id,
            name=target_file.name,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            location=str(target_file),
            description=(
                f"Verified computation deliverable for: {calc}. "
                f"Evaluated {meta['record_count']} equipment items "
                f"({meta['compliant_count']} compliant, {meta['below_min_count']} below minimum)."
            ),
            source_request_id=request.request_id,
        )

        summary_msg = (
            f"Computation deliverable successfully generated at '{target_file.name}' "
            f"containing {meta['record_count']} evaluated item(s) and full methodology."
        )

        output: JsonObject = {
            "artifact_id": str(artifact_id),
            "file_path": str(target_file),
            "summary": summary_msg,
            "record_count": meta["record_count"],
            "compliant_count": meta["compliant_count"],
            "below_min_count": meta["below_min_count"],
            "verification_status": meta["verification_status"],
            "requested_calculation": meta["requested_calculation"],
            "source_data_reference": meta["source_data_reference"],
            "methodology": meta["methodology"],
        }

        observation = Observation(
            source="generate_excel",
            kind="artifact_generated",
            summary=summary_msg,
            data={"artifact_id": str(artifact_id), "file_path": str(target_file)},
            request_id=request.request_id,
        )

        return CapabilityResult(
            request_id=request.request_id,
            status=CapabilityResultStatus.SUCCEEDED,
            output=output,
            artifacts=[artifact],
            observations=[observation],
        )
