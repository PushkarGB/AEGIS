"""Tests for deterministic spreadsheet inspection using openpyxl."""

from __future__ import annotations

import datetime
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from aegis.broker import RegistryCapabilityBroker
from aegis.capabilities import (
    CapabilityKind,
    CapabilityRegistry,
    ColumnInfo,
    InspectSpreadsheetCapability,
    SheetInfo,
    WorkbookInspection,
    WorkbookMetadata,
    inspect_spreadsheet,
)
from aegis.config import load_config
from aegis.orchestration import ExecutionController, WorkflowName
from aegis.schemas import (
    AgentDecision,
    CapabilityRequest,
    CapabilityResultStatus,
    TaskState,
)


@pytest.fixture
def synthetic_equipment_workbook(tmp_path: Path) -> Path:
    """Create a multi-sheet synthetic Excel workbook representing equipment readings."""
    wb = openpyxl.Workbook()

    # Set document properties
    wb.properties.title = "Plant Equipment Thickness Inspection"
    wb.properties.creator = "AEGIS Quality Control"

    # Sheet 1: Main equipment readings
    ws1 = wb.active
    ws1.title = "EquipmentReadings"

    headers1 = [
        "equipment_id",
        "inspection_date",
        "measured_thickness",
        "min_thickness",
        "cycle_count",
        "is_active",
        "inspector_notes",
    ]
    ws1.append(headers1)

    rows1 = [
        ["EQ-101", datetime.date(2026, 8, 1), 12.5, 10.0, 1500, True, "Normal condition"],
        ["EQ-102", datetime.date(2026, 8, 2), 8.2, 9.0, 3200, True, "Below minimum thickness threshold"],
        ["EQ-103", datetime.date(2026, 8, 3), 15.0, 12.0, 800, True, None],
        ["EQ-104", datetime.date(2026, 8, 4), 6.8, 8.5, 4100, False, "Scheduled for replacement"],
        ["EQ-105", datetime.date(2026, 8, 5), 11.4, 10.0, 2100, True, "Minor wear"],
        ["EQ-106", datetime.date(2026, 8, 6), 9.9, 10.0, 2900, True, "Borderline thickness"],
        ["EQ-107", datetime.date(2026, 8, 7), 14.2, 11.0, 1100, True, "Good"],
    ]
    for row in rows1:
        ws1.append(row)

    # Sheet 2: Summary metrics
    ws2 = wb.create_sheet(title="SummaryMetrics")
    headers2 = ["section_name", "total_equipment", "avg_thickness", "status"]
    ws2.append(headers2)
    ws2.append(["Boiler Section A", 4, 10.625, "PASS"])
    ws2.append(["Piping Section B", 3, 11.833, "REVIEW"])

    # Sheet 3: Completely empty sheet
    wb.create_sheet(title="EmptySheet")

    file_path = tmp_path / "equipment_readings.xlsx"
    wb.save(str(file_path))
    wb.close()
    return file_path


@pytest.fixture
def synthetic_edge_case_workbook(tmp_path: Path) -> Path:
    """Create a workbook with edge cases: unnamed columns, duplicate headers, header-only."""
    wb = openpyxl.Workbook()

    # Sheet 1: Duplicate headers and missing header cell
    ws1 = wb.active
    ws1.title = "EdgeCases"
    ws1.append(["Reading", "Reading", "", "Value"])
    ws1.append([10.5, 20.5, "extra", 100])
    ws1.append([12.0, 22.0, "notes", 200])

    # Sheet 2: Header-only sheet (no data rows)
    ws2 = wb.create_sheet(title="HeaderOnly")
    ws2.append(["id", "name", "count"])

    file_path = tmp_path / "edge_cases.xlsx"
    wb.save(str(file_path))
    wb.close()
    return file_path


# ── Direct Function Tests ─────────────────────────────────────────────


def test_inspect_spreadsheet_returns_structured_workbook_info(synthetic_equipment_workbook: Path):
    inspection = inspect_spreadsheet(synthetic_equipment_workbook)

    assert isinstance(inspection, WorkbookInspection)
    assert inspection.total_sheets == 3
    assert inspection.sheet_names == ["EquipmentReadings", "SummaryMetrics", "EmptySheet"]

    # Basic metadata
    assert inspection.metadata.file_name == "equipment_readings.xlsx"
    assert inspection.metadata.file_size_bytes > 0
    assert inspection.metadata.title == "Plant Equipment Thickness Inspection"
    assert inspection.metadata.creator == "AEGIS Quality Control"
    assert inspection.metadata.active_sheet == "EquipmentReadings"

    # Primary sheet convenience fields
    assert inspection.columns == [
        "equipment_id",
        "inspection_date",
        "measured_thickness",
        "min_thickness",
        "cycle_count",
        "is_active",
        "inspector_notes",
    ]
    assert inspection.row_count == 7
    assert inspection.numeric_fields == ["measured_thickness", "min_thickness", "cycle_count"]
    assert len(inspection.preview_rows) == 5  # default max preview rows


def test_inspect_spreadsheet_column_details_and_type_inference(synthetic_equipment_workbook: Path):
    inspection = inspect_spreadsheet(synthetic_equipment_workbook)
    sheet1 = inspection.sheets[0]

    assert sheet1.name == "EquipmentReadings"
    assert sheet1.is_active is True
    assert sheet1.is_empty is False
    assert sheet1.data_row_count == 7
    assert sheet1.column_count == 7

    col_map: dict[str, ColumnInfo] = {c.name: c for c in sheet1.column_details}

    # equipment_id: string
    eq_col = col_map["equipment_id"]
    assert eq_col.letter == "A"
    assert eq_col.index == 1
    assert eq_col.inferred_type == "string"
    assert eq_col.is_numeric is False
    assert eq_col.non_empty_count == 7
    assert eq_col.null_count == 0
    assert "EQ-101" in eq_col.sample_values

    # inspection_date: datetime / date
    date_col = col_map["inspection_date"]
    assert date_col.letter == "B"
    assert date_col.inferred_type == "datetime"
    assert date_col.is_numeric is False
    assert date_col.sample_values[0].startswith("2026-08-01")

    # measured_thickness: float (numeric)
    thick_col = col_map["measured_thickness"]
    assert thick_col.letter == "C"
    assert thick_col.inferred_type == "float"
    assert thick_col.is_numeric is True
    assert thick_col.min_value == 6.8
    assert thick_col.max_value == 15.0
    assert thick_col.sample_values == [12.5, 8.2, 15.0, 6.8, 11.4]

    # cycle_count: integer (numeric)
    cycle_col = col_map["cycle_count"]
    assert cycle_col.letter == "E"
    assert cycle_col.inferred_type == "integer"
    assert cycle_col.is_numeric is True
    assert cycle_col.min_value == 800
    assert cycle_col.max_value == 4100

    # is_active: boolean
    active_col = col_map["is_active"]
    assert active_col.letter == "F"
    assert active_col.inferred_type == "boolean"
    assert active_col.is_numeric is False

    # inspector_notes: string with nulls
    notes_col = col_map["inspector_notes"]
    assert notes_col.letter == "G"
    assert notes_col.inferred_type == "string"
    assert notes_col.null_count == 1
    assert notes_col.non_empty_count == 6


def test_inspect_spreadsheet_multiple_sheets_and_empty_sheet(synthetic_equipment_workbook: Path):
    inspection = inspect_spreadsheet(synthetic_equipment_workbook)

    # Sheet 2: SummaryMetrics
    sheet2 = inspection.sheets[1]
    assert sheet2.name == "SummaryMetrics"
    assert sheet2.is_empty is False
    assert sheet2.data_row_count == 2
    assert sheet2.columns == ["section_name", "total_equipment", "avg_thickness", "status"]
    assert sheet2.numeric_fields == ["total_equipment", "avg_thickness"]

    # Sheet 3: EmptySheet
    sheet3 = inspection.sheets[2]
    assert sheet3.name == "EmptySheet"
    assert sheet3.is_empty is True
    assert sheet3.data_row_count == 0
    assert sheet3.total_rows == 0
    assert sheet3.columns == []
    assert sheet3.column_details == []


def test_inspect_spreadsheet_edge_cases(synthetic_edge_case_workbook: Path):
    inspection = inspect_spreadsheet(synthetic_edge_case_workbook)

    sheet1 = inspection.sheets[0]
    # Header deduplication and blank column naming
    assert sheet1.columns == ["Reading", "Reading_1", "column_C", "Value"]
    assert sheet1.data_row_count == 2
    assert sheet1.numeric_fields == ["Reading", "Reading_1", "Value"]

    sheet2 = inspection.sheets[1]
    # Header only sheet
    assert sheet2.name == "HeaderOnly"
    assert sheet2.is_empty is False
    assert sheet2.data_row_count == 0
    assert sheet2.columns == ["id", "name", "count"]


def test_inspect_spreadsheet_file_not_found(tmp_path: Path):
    non_existent = tmp_path / "does_not_exist.xlsx"
    with pytest.raises(FileNotFoundError, match="Spreadsheet file not found"):
        inspect_spreadsheet(non_existent)


def test_inspect_spreadsheet_invalid_file(tmp_path: Path):
    bad_file = tmp_path / "not_an_excel.xlsx"
    bad_file.write_text("Hello, this is not an Excel binary file.")
    with pytest.raises(ValueError, match="Failed to open workbook"):
        inspect_spreadsheet(bad_file)


# ── Capability and Broker Tests ───────────────────────────────────────


def test_capability_metadata_and_structure():
    cap = InspectSpreadsheetCapability()
    meta = cap.metadata

    assert meta.name == "inspect_spreadsheet"
    assert meta.kind == CapabilityKind.TOOL
    assert meta.input_modalities == ("spreadsheet",)
    assert "openpyxl" in meta.description or "inspect" in meta.description.lower()
    assert "required" in meta.output_contract.json_schema


def test_capability_executes_successfully_via_request(synthetic_equipment_workbook: Path):
    cap = InspectSpreadsheetCapability()
    request = CapabilityRequest(
        capability_name="inspect_spreadsheet",
        inputs={"workbook": str(synthetic_equipment_workbook)},
    )

    result = cap.invoke(request)

    assert result.status == CapabilityResultStatus.SUCCEEDED
    assert result.error is None
    assert result.output["total_sheets"] == 3
    assert result.output["row_count"] == 7
    assert result.output["columns"] == [
        "equipment_id",
        "inspection_date",
        "measured_thickness",
        "min_thickness",
        "cycle_count",
        "is_active",
        "inspector_notes",
    ]
    assert result.output["numeric_fields"] == ["measured_thickness", "min_thickness", "cycle_count"]

    # Check generated observation
    assert len(result.observations) == 1
    obs = result.observations[0]
    assert obs.source == "inspect_spreadsheet"
    assert obs.kind == "spreadsheet_inspection"
    assert "equipment_readings.xlsx" in obs.summary
    assert obs.data["columns"] == result.output["columns"]
    assert obs.data["row_count"] == 7


def test_capability_supports_alternative_input_keys(synthetic_equipment_workbook: Path):
    cap = InspectSpreadsheetCapability()

    for key in ("file_path", "path", "filepath", "attachment"):
        request = CapabilityRequest(
            capability_name="inspect_spreadsheet",
            inputs={key: str(synthetic_equipment_workbook)},
        )
        result = cap.invoke(request)
        assert result.status == CapabilityResultStatus.SUCCEEDED


def test_capability_fails_cleanly_on_missing_or_invalid_inputs(tmp_path: Path):
    cap = InspectSpreadsheetCapability()

    # Missing path in inputs
    no_input_req = CapabilityRequest(
        capability_name="inspect_spreadsheet",
        inputs={},
    )
    res1 = cap.invoke(no_input_req)
    assert res1.status == CapabilityResultStatus.FAILED
    assert "No workbook file path provided" in res1.error

    # Non-existent file
    missing_file_req = CapabilityRequest(
        capability_name="inspect_spreadsheet",
        inputs={"workbook": str(tmp_path / "ghost.xlsx")},
    )
    res2 = cap.invoke(missing_file_req)
    assert res2.status == CapabilityResultStatus.FAILED
    assert "not found" in res2.error


def test_capability_registered_in_registry_and_invoked_by_broker(synthetic_equipment_workbook: Path):
    registry = CapabilityRegistry(load_config().capabilities)
    real_capability = InspectSpreadsheetCapability()
    registry.register(real_capability)

    broker = RegistryCapabilityBroker(registry)
    request = CapabilityRequest(
        capability_name="inspect_spreadsheet",
        inputs={"workbook": str(synthetic_equipment_workbook)},
    )

    result = broker.invoke(request)

    assert result.status == CapabilityResultStatus.SUCCEEDED
    assert result.output["row_count"] == 7


def test_capability_integrates_with_execution_controller(synthetic_equipment_workbook: Path):
    registry = CapabilityRegistry(load_config().capabilities)
    registry.register(InspectSpreadsheetCapability())
    broker = RegistryCapabilityBroker(registry)

    state = TaskState(
        user_goal="Calculate average equipment thickness and find minimum violations.",
        attachments=[str(synthetic_equipment_workbook)],
    )
    controller = ExecutionController(state, WorkflowName.COMPUTATION, broker)

    # Controller executes first step of computation workflow
    controller.execute(
        AgentDecision(
            action="inspect_spreadsheet",
            inputs={"workbook": str(synthetic_equipment_workbook)},
        )
    )

    assert state.current_step == "generate"
    assert "inspect_spreadsheet" in state.completed_steps
    assert len(state.observations) == 2  # inspection observation + controller capability_succeeded observation
    assert state.observations[0].source == "inspect_spreadsheet"
    assert state.observations[0].kind == "spreadsheet_inspection"
    assert state.observations[0].data["row_count"] == 7
