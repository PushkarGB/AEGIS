"""Tests for generate_code integration with ModelRouter and ModelProvider.

Verifies:
1. Integration of generate_code with ModelRouter and ModelProvider:
   - Routes to the coding model role deterministically.
   - Provider receives correctly formed ModelGenerationRequest.
   - Works with single ModelProvider or provider dictionary mapping.
2. The Coding Model receives:
   - computation objective;
   - relevant spreadsheet structure/data description;
   - required output constraints.
3. Returning executable Python code:
   - Returns valid Python code (verifiable with ast.parse).
   - Strips markdown code fences (```python ... ```, ``` ... ```).
   - Handles text commentary around code blocks.
4. Non-execution invariant:
   - Does NOT execute the generated code in any way.
5. Error handling and validation:
   - Missing/empty computation objective.
   - Missing provider mapping.
   - Router failures.
   - Empty model response.
6. Standalone function generate_code(...) and GenerateCodeCapability.
"""

from __future__ import annotations

import ast
import json
from uuid import uuid4

import pytest

from aegis.capabilities import (
    CapabilityKind,
    CapabilityRegistry,
    GenerateCodeCapability,
    generate_code,
)
from aegis.config import load_config
from aegis.router import (
    MockModelProvider,
    ModelGenerationRequest,
    ModelRegistry,
    ModelRouter,
    RoutingDecision,
    RoutingError,
)
from aegis.schemas import (
    CapabilityRequest,
    CapabilityResultStatus,
)


def _setup_router_and_mock_provider(response_text: str = "print('calculated')"):
    """Helper to set up ModelRegistry, ModelRouter, and MockModelProvider."""
    config = load_config()
    registry = ModelRegistry(config.models)
    router = ModelRouter(registry)

    coding_models = registry.get_models_for_role("coding")
    assert coding_models, "No coding model configured"
    coding_model = coding_models[0]

    captured_requests: list[ModelGenerationRequest] = []

    def mock_generate(req: ModelGenerationRequest) -> str:
        captured_requests.append(req)
        return response_text

    provider = MockModelProvider(response_factory=mock_generate)
    providers = {coding_model.provider: provider}

    return router, provider, providers, captured_requests, coding_model


class TestModelRouterAndProviderIntegration:
    """Verify generate_code cleanly integrates with ModelRouter and ModelProvider."""

    def test_routes_to_coding_model_role(self):
        router, provider, providers, captured, coding_model = _setup_router_and_mock_provider(
            "import openpyxl\nresult = 42\nprint(result)"
        )
        cap = GenerateCodeCapability(router=router, providers=providers)

        req = CapabilityRequest(
            capability_name="generate_code",
            inputs={
                "computation_objective": "Compute average thickness per equipment",
                "relevant_spreadsheet_structure": "Columns: Equipment_ID, Thickness",
                "required_output_constraints": ["Print results to stdout"],
            },
        )
        result = cap.invoke(req)

        assert result.status == CapabilityResultStatus.SUCCEEDED
        assert result.output["model_id"] == coding_model.id
        assert len(captured) == 1
        assert captured[0].model_id == coding_model.id

    def test_works_with_single_model_provider_instance(self):
        router, provider, _, captured, coding_model = _setup_router_and_mock_provider(
            "print('single provider works')"
        )
        # Pass single provider directly
        cap = GenerateCodeCapability(router=router, provider=provider)

        req = CapabilityRequest(
            capability_name="generate_code",
            inputs={
                "computation_objective": "Find max pressure",
                "spreadsheet_structure": "Sheet: Readings, Col: Pressure",
            },
        )
        result = cap.invoke(req)

        assert result.status == CapabilityResultStatus.SUCCEEDED
        assert "print('single provider works')" in result.output["code"]
        assert len(captured) == 1

    def test_fails_when_provider_not_found(self):
        router, _, _, _, _ = _setup_router_and_mock_provider()
        cap = GenerateCodeCapability(router=router, providers={})

        req = CapabilityRequest(
            capability_name="generate_code",
            inputs={"computation_objective": "Calculate something"},
        )
        result = cap.invoke(req)

        assert result.status == CapabilityResultStatus.FAILED
        assert "no provider" in result.error.lower()

    def test_observation_contains_routing_metadata(self):
        router, provider, providers, _, coding_model = _setup_router_and_mock_provider("x = 1")
        cap = GenerateCodeCapability(router=router, providers=providers)

        req = CapabilityRequest(
            capability_name="generate_code",
            inputs={"computation_objective": "Compute total"},
        )
        result = cap.invoke(req)

        assert result.status == CapabilityResultStatus.SUCCEEDED
        assert len(result.observations) == 1
        obs = result.observations[0]
        assert obs.source == "generate_code"
        assert obs.kind == "code_generated"
        assert obs.data["model_id"] == coding_model.id
        assert obs.data["provider_id"] == coding_model.provider


class TestCodingModelPromptInputs:
    """Verify the Coding Model receives all three required pieces of context:
    - computation objective;
    - relevant spreadsheet structure/data description;
    - required output constraints.
    """

    def test_coding_model_receives_objective_structure_and_constraints(self):
        router, provider, providers, captured, _ = _setup_router_and_mock_provider(
            "import openpyxl\nprint('done')"
        )
        cap = GenerateCodeCapability(router=router, providers=providers)

        objective = "Identify equipment with thickness below minimum acceptable threshold"
        structure = "Columns: [Equipment_ID (str), Measured_Thickness (float), Min_Acceptable (float)]"
        constraints = [
            "Use openpyxl to read data",
            "Print comma-separated list of failing equipment IDs",
            "Do not write output files",
        ]

        req = CapabilityRequest(
            capability_name="generate_code",
            inputs={
                "computation_objective": objective,
                "relevant_spreadsheet_structure": structure,
                "required_output_constraints": constraints,
                "file_path": "/data/inspection.xlsx",
            },
        )
        result = cap.invoke(req)

        assert result.status == CapabilityResultStatus.SUCCEEDED
        assert len(captured) == 1
        prompt = captured[0].prompt

        # 1. Coding model receives computation objective
        assert "Computation Objective:" in prompt
        assert objective in prompt

        # 2. Coding model receives relevant spreadsheet structure/data description
        assert "Relevant Spreadsheet Structure / Data Description:" in prompt
        assert structure in prompt

        # 3. Coding model receives required output constraints
        assert "Required Output Constraints:" in prompt
        for c in constraints:
            assert c in prompt

        # File path is included
        assert "/data/inspection.xlsx" in prompt

    def test_coding_model_receives_structured_dict_as_spreadsheet_structure(self):
        router, provider, providers, captured, _ = _setup_router_and_mock_provider("print(1)")
        cap = GenerateCodeCapability(router=router, providers=providers)

        dict_structure = {
            "sheet_names": ["Measurements"],
            "columns": ["Tag", "Flow_Rate", "Temperature"],
            "row_count": 250,
            "numeric_fields": ["Flow_Rate", "Temperature"],
        }

        req = CapabilityRequest(
            capability_name="generate_code",
            inputs={
                "computation_objective": "Average flow rate",
                "spreadsheet_structure": dict_structure,
            },
        )
        result = cap.invoke(req)

        assert result.status == CapabilityResultStatus.SUCCEEDED
        prompt = captured[0].prompt
        assert "Measurements" in prompt
        assert "Flow_Rate" in prompt
        assert "250" in prompt

    def test_coding_model_receives_default_output_constraints_when_none_provided(self):
        router, provider, providers, captured, _ = _setup_router_and_mock_provider("print(1)")
        cap = GenerateCodeCapability(router=router, providers=providers)

        req = CapabilityRequest(
            capability_name="generate_code",
            inputs={"computation_objective": "Sum column B"},
        )
        result = cap.invoke(req)

        assert result.status == CapabilityResultStatus.SUCCEEDED
        prompt = captured[0].prompt
        # Default output constraints must be present
        assert "Required Output Constraints:" in prompt
        assert "openpyxl" in prompt
        assert "stdout" in prompt
        assert "network" in prompt.lower()

    def test_coding_model_receives_correction_context_on_retry(self):
        router, provider, providers, captured, _ = _setup_router_and_mock_provider("print('fixed')")
        cap = GenerateCodeCapability(router=router, providers=providers)

        req = CapabilityRequest(
            capability_name="generate_code",
            inputs={
                "computation_objective": "Calculate yield",
                "spreadsheet_structure": "Cols: A, B",
                "correction_context": "Previous run crashed: ZeroDivisionError on line 5",
            },
        )
        result = cap.invoke(req)

        assert result.status == CapabilityResultStatus.SUCCEEDED
        prompt = captured[0].prompt
        assert "Correction Context" in prompt
        assert "ZeroDivisionError" in prompt

    def test_supports_flexible_input_aliases(self):
        router, provider, providers, captured, _ = _setup_router_and_mock_provider("print(1)")
        cap = GenerateCodeCapability(router=router, providers=providers)

        # Using alternative key names: computation, data_schema, constraints
        req = CapabilityRequest(
            capability_name="generate_code",
            inputs={
                "computation": "Calculate total volume",
                "data_schema": "V1, V2, V3",
                "constraints": ["Must be deterministic"],
            },
        )
        result = cap.invoke(req)

        assert result.status == CapabilityResultStatus.SUCCEEDED
        prompt = captured[0].prompt
        assert "Calculate total volume" in prompt
        assert "V1, V2, V3" in prompt
        assert "Must be deterministic" in prompt


class TestExecutablePythonCodeReturn:
    """Verify generate_code returns clean, executable Python code."""

    def test_returns_raw_executable_python_code(self):
        raw_code = (
            "import openpyxl\n\n"
            "wb = openpyxl.load_workbook('data.xlsx')\n"
            "sheet = wb.active\n"
            "total = sum(row[1].value for row in sheet.iter_rows(min_row=2))\n"
            "print(f'Total: {total}')\n"
        )
        router, provider, providers, _, _ = _setup_router_and_mock_provider(raw_code)
        cap = GenerateCodeCapability(router=router, providers=providers)

        req = CapabilityRequest(
            capability_name="generate_code",
            inputs={"computation_objective": "Compute sum"},
        )
        result = cap.invoke(req)

        assert result.status == CapabilityResultStatus.SUCCEEDED
        code = result.output["code"]
        assert code == raw_code.strip()
        # Verify valid Python syntax using ast.parse
        parsed = ast.parse(code)
        assert isinstance(parsed, ast.Module)

    def test_strips_markdown_python_code_fences(self):
        fenced_code = (
            "```python\n"
            "import openpyxl\n"
            "print('Hello from fenced block')\n"
            "```"
        )
        router, provider, providers, _, _ = _setup_router_and_mock_provider(fenced_code)
        cap = GenerateCodeCapability(router=router, providers=providers)

        req = CapabilityRequest(
            capability_name="generate_code",
            inputs={"computation_objective": "Test fences"},
        )
        result = cap.invoke(req)

        assert result.status == CapabilityResultStatus.SUCCEEDED
        code = result.output["code"]
        assert "```" not in code
        assert "import openpyxl" in code
        ast.parse(code)

    def test_strips_commentary_around_code_fence(self):
        messy_response = (
            "Here is the Python script to compute the values:\n\n"
            "```python\n"
            "import openpyxl\n"
            "wb = openpyxl.Workbook()\n"
            "print('Executed successfully')\n"
            "```\n\n"
            "This script reads the sheet and prints the outcome."
        )
        router, provider, providers, _, _ = _setup_router_and_mock_provider(messy_response)
        cap = GenerateCodeCapability(router=router, providers=providers)

        req = CapabilityRequest(
            capability_name="generate_code",
            inputs={"computation_objective": "Compute values"},
        )
        result = cap.invoke(req)

        assert result.status == CapabilityResultStatus.SUCCEEDED
        code = result.output["code"]
        assert "Here is the Python script" not in code
        assert "This script reads the sheet" not in code
        assert "import openpyxl" in code
        ast.parse(code)

    def test_fails_on_empty_model_response(self):
        router, provider, providers, _, _ = _setup_router_and_mock_provider("   \n  ")
        cap = GenerateCodeCapability(router=router, providers=providers)

        req = CapabilityRequest(
            capability_name="generate_code",
            inputs={"computation_objective": "Compute values"},
        )
        result = cap.invoke(req)

        assert result.status == CapabilityResultStatus.FAILED
        assert "empty code" in result.error.lower()


class TestNonExecutionInvariant:
    """Verify that generate_code NEVER executes the generated code."""

    def test_does_not_execute_malicious_or_side_effect_code(self, tmp_path):
        canary_file = tmp_path / "should_never_be_created.txt"
        dangerous_code = (
            f"with open(r'{canary_file}', 'w') as f:\n"
            f"    f.write('I executed!')\n"
        )
        router, provider, providers, _, _ = _setup_router_and_mock_provider(dangerous_code)
        cap = GenerateCodeCapability(router=router, providers=providers)

        req = CapabilityRequest(
            capability_name="generate_code",
            inputs={"computation_objective": "Dangerous operation"},
        )
        result = cap.invoke(req)

        assert result.status == CapabilityResultStatus.SUCCEEDED
        assert "should_never_be_created" in result.output["code"]
        # Invariant: code must NOT have been executed!
        assert not canary_file.exists(), "Code was executed during generate_code! Architectural invariant violated."


class TestStandaloneGenerateCodeFunction:
    """Verify the standalone generate_code(...) function."""

    def test_standalone_generate_code_function(self):
        code_to_return = "import openpyxl\nprint('standalone works')\n"
        router, provider, _, captured, _ = _setup_router_and_mock_provider(code_to_return)

        result_code = generate_code(
            router=router,
            provider=provider,
            computation_objective="Calculate corrosion rate",
            spreadsheet_structure="Sheet1: Date, Rate",
            output_constraints=["Print rate"],
            file_path="/tmp/corrosion.xlsx",
        )

        assert result_code == code_to_return.strip()
        assert len(captured) == 1
        prompt = captured[0].prompt
        assert "Calculate corrosion rate" in prompt
        assert "Sheet1: Date, Rate" in prompt
        assert "Print rate" in prompt
        assert "/tmp/corrosion.xlsx" in prompt

    def test_standalone_rejects_empty_objective(self):
        router, provider, _, _, _ = _setup_router_and_mock_provider("print(1)")

        with pytest.raises(ValueError, match="computation_objective"):
            generate_code(
                router=router,
                provider=provider,
                computation_objective="",
            )

    def test_standalone_fails_on_missing_provider(self):
        router, _, _, _, _ = _setup_router_and_mock_provider("print(1)")

        with pytest.raises(ValueError, match="No ModelProvider"):
            generate_code(
                router=router,
                provider={},
                computation_objective="Test",
            )
