"""Comprehensive tests for computation deliverable generation (generate_excel capability).

Verifies:
1. Deliverable artifact generation containing:
   - requested calculation;
   - source data reference;
   - result (summary metrics and detailed items);
   - relevant methodology;
   - verification status.
2. Appropriate local spreadsheet format (.xlsx via openpyxl with multiple sheets).
3. Number formatting and compliance status visualization.
4. Flexible input resolution across alias keys.
5. ExecutionController integration in COMPUTATION_WORKFLOW (deliver -> finish).
6. Separation / sovereign invariant: no model chain-of-thought exposed.
"""

from __future__ import annotations

import json
from pathlib import Path
from tempfile import TemporaryDirectory

import openpyxl
import pytest

from aegis.broker import RegistryCapabilityBroker
from aegis.capabilities import (
    CapabilityKind,
    CapabilityRegistry,
    GenerateExcelCapability,
    generate_excel_deliverable,
)
from aegis.config import load_config
from aegis.orchestration import ExecutionController, ExecutionEventKind, WorkflowName
from aegis.schemas import (
    AgentDecision,
    CapabilityRequest,
    CapabilityResult,
    CapabilityResultStatus,
    FinalStatus,
    TaskState,
    VerificationStatus,
)


# ---------------------------------------------------------------------------
# 1. Standalone Deliverable Generation Tests
# ---------------------------------------------------------------------------


class TestGenerateExcelDeliverableFunction:
    """Verify generate_excel_deliverable produces a structured, formatted workbook."""

    def test_deliverable_contains_all_five_required_elements(self):
        with TemporaryDirectory() as tmpdir:
            sample_records = [
                {
                    "equipment_id": "EQ-001",
                    "average_measured_thickness": 4.52,
                    "min_acceptable_thickness": 3.0,
                    "below_min": False,
                },
                {
                    "equipment_id": "EQ-002",
                    "average_measured_thickness": 2.74,
                    "min_acceptable_thickness": 3.0,
                    "below_min": True,
                },
            ]

            target_file, meta = generate_excel_deliverable(
                requested_calculation="Calculate average measured thickness and flag items below minimum",
                source_data_reference="/data/inspection_readings.xlsx",
                result_data=sample_records,
                methodology="Arithmetic mean of measured thickness per equipment compared to threshold",
                verification_status="VERIFIED — All 4 deterministic checks passed",
                output_dir=tmpdir,
                filename="test_deliverable.xlsx",
            )

            assert target_file.exists()
            assert target_file.is_file()

            # Verify returned metadata
            assert meta["requested_calculation"] == "Calculate average measured thickness and flag items below minimum"
            assert meta["source_data_reference"] == "/data/inspection_readings.xlsx"
            assert meta["methodology"] == "Arithmetic mean of measured thickness per equipment compared to threshold"
            assert meta["verification_status"] == "VERIFIED — All 4 deterministic checks passed"
            assert meta["record_count"] == 2
            assert meta["compliant_count"] == 1
            assert meta["below_min_count"] == 1

            # Inspect actual Excel workbook contents
            wb = openpyxl.load_workbook(target_file, data_only=True)
            assert "Calculation Summary" in wb.sheetnames
            assert "Detailed Results" in wb.sheetnames

            ws_summary = wb["Calculation Summary"]
            # 1. Requested calculation
            assert ws_summary["B4"].value == "Calculate average measured thickness and flag items below minimum"
            # 2. Source data reference
            assert ws_summary["B5"].value == "/data/inspection_readings.xlsx"
            # 3. Relevant methodology
            assert ws_summary["B6"].value == "Arithmetic mean of measured thickness per equipment compared to threshold"
            # 4. Verification status
            assert ws_summary["B7"].value == "VERIFIED — All 4 deterministic checks passed"

            # 5. Results (KPI summary)
            summary_text = [ws_summary.cell(row=r, column=1).value for r in range(4, 20)]
            assert any("Total Equipment Evaluated" in str(txt) for txt in summary_text)
            assert any("Compliant" in str(txt) for txt in summary_text)
            assert any("Below Minimum" in str(txt) for txt in summary_text)

            # Detailed results sheet
            ws_details = wb["Detailed Results"]
            headers = [ws_details.cell(row=1, column=c).value for c in range(1, 5)]
            assert "Equipment Id" in headers
            assert "Average Measured Thickness" in headers

            # Check rows
            eq_ids = [ws_details.cell(row=r, column=1).value for r in (2, 3)]
            assert "EQ-001" in eq_ids
            assert "EQ-002" in eq_ids

    def test_delivers_from_raw_json_stdout(self):
        with TemporaryDirectory() as tmpdir:
            json_stdout = json.dumps([
                {"equipment_id": "EQ-001", "average": 4.5, "below_min": False},
                {"equipment_id": "EQ-003", "average": 2.1, "below_min": True},
            ])

            target_file, meta = generate_excel_deliverable(
                requested_calculation="Average thickness",
                source_data_reference="/data/readings.xlsx",
                stdout=json_stdout,
                output_dir=tmpdir,
            )

            assert target_file.exists()
            assert meta["record_count"] == 2
            assert meta["compliant_count"] == 1
            assert meta["below_min_count"] == 1

            wb = openpyxl.load_workbook(target_file, data_only=True)
            ws_details = wb["Detailed Results"]
            assert ws_details.cell(row=2, column=1).value == "EQ-001"

    def test_formatting_of_numbers_and_compliance_indicators(self):
        with TemporaryDirectory() as tmpdir:
            records = [
                {"equipment_id": "EQ-001", "average": 4.5238, "below_min": False},
                {"equipment_id": "EQ-002", "average": 2.7000, "below_min": True},
            ]

            target_file, _ = generate_excel_deliverable(
                requested_calculation="Calc",
                source_data_reference="data.xlsx",
                result_data=records,
                output_dir=tmpdir,
            )

            wb = openpyxl.load_workbook(target_file)
            ws_details = wb["Detailed Results"]

            # Number formatting check
            avg_cell = ws_details.cell(row=2, column=2)
            assert avg_cell.number_format == "0.00"

            # Status cell formatting check
            status_cell_1 = ws_details.cell(row=2, column=3)
            status_cell_2 = ws_details.cell(row=3, column=3)

            assert status_cell_1.value == "COMPLIANT"
            assert status_cell_2.value == "BELOW MINIMUM"


# ---------------------------------------------------------------------------
# 2. GenerateExcelCapability Tests
# ---------------------------------------------------------------------------


class TestGenerateExcelCapability:
    """Verify GenerateExcelCapability tool execution, metadata, and contracts."""

    def test_capability_metadata(self):
        cap = GenerateExcelCapability()
        assert cap.metadata.name == "generate_excel"
        assert cap.metadata.kind == CapabilityKind.TOOL
        assert "spreadsheet" in cap.metadata.input_modalities

    def test_capability_execution_produces_artifact_and_observation(self):
        with TemporaryDirectory() as tmpdir:
            cap = GenerateExcelCapability(output_dir=tmpdir)
            req = CapabilityRequest(
                capability_name="generate_excel",
                inputs={
                    "requested_calculation": "Equipment thickness evaluation",
                    "source_data_reference": "inspection_readings.xlsx",
                    "result": [{"equipment_id": "EQ-001", "average": 4.5}],
                    "methodology": "Mean calculation",
                    "verification_status": "VERIFIED",
                },
            )

            result = cap.invoke(req)

            assert result.status == CapabilityResultStatus.SUCCEEDED
            assert len(result.artifacts) == 1
            artifact = result.artifacts[0]
            assert artifact.name.endswith(".xlsx")
            assert artifact.media_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            assert Path(artifact.location).exists()

            assert len(result.observations) == 1
            obs = result.observations[0]
            assert obs.source == "generate_excel"
            assert obs.kind == "artifact_generated"
            assert "deliverable successfully generated" in obs.summary

            assert result.output["record_count"] == 1
            assert result.output["verification_status"] == "VERIFIED"

    def test_resolves_flexible_alias_keys(self):
        with TemporaryDirectory() as tmpdir:
            cap = GenerateExcelCapability(output_dir=tmpdir)
            # Using computation_objective, file_path, data
            req = CapabilityRequest(
                capability_name="generate_excel",
                inputs={
                    "computation_objective": "Average Flow Rate",
                    "file_path": "/data/flow.xlsx",
                    "data": [{"tag": "F-101", "average_flow": 250.5}],
                },
            )

            result = cap.invoke(req)
            assert result.status == CapabilityResultStatus.SUCCEEDED
            assert result.output["requested_calculation"] == "Average Flow Rate"
            assert result.output["source_data_reference"] == "/data/flow.xlsx"
            assert result.output["record_count"] == 1


# ---------------------------------------------------------------------------
# 3. ExecutionController Workflow B Integration
# ---------------------------------------------------------------------------


class TestControllerComputationWorkflowIntegration:
    """Verify ExecutionController transitions through the deliver state to finish."""

    def test_controller_executes_generate_excel_and_transitions_to_finish(self):
        with TemporaryDirectory() as tmpdir:
            config = load_config()
            registry = CapabilityRegistry(config.capabilities)

            from aegis.capabilities.base import Capability, CapabilityMetadata

            class _MockFinish(Capability):
                @property
                def metadata(self):
                    return CapabilityMetadata(
                        name="finish",
                        kind=CapabilityKind.CONTROL,
                        description="Mock finish.",
                        input_modalities=("spreadsheet", "document", "image"),
                    )

                def execute(self, request):
                    return CapabilityResult(
                        request_id=request.request_id,
                        status=CapabilityResultStatus.SUCCEEDED,
                    )

            excel_cap = GenerateExcelCapability(output_dir=tmpdir)
            registry.register(excel_cap)
            registry.register(_MockFinish())
            broker = RegistryCapabilityBroker(registry)

            state = TaskState(
                user_goal="Calculate average thickness per equipment item.",
                current_step="deliver",
                verification_status=VerificationStatus.PASSED,
            )
            controller = ExecutionController(state, WorkflowName.COMPUTATION, broker)

            decision = AgentDecision(
                action="generate_excel",
                inputs={
                    "requested_calculation": "Average thickness calculation",
                    "source_data_reference": "readings.xlsx",
                    "result": [{"equipment_id": "EQ-001", "average": 4.5}],
                },
            )
            event = controller.execute(decision)

            assert event.kind == ExecutionEventKind.ACTION_COMPLETED
            assert event.action == "generate_excel"
            assert controller.state.current_step == "finish"
            assert "generate_excel" in controller.state.completed_steps
            assert len(controller.state.generated_artifacts) == 1
            assert controller.state.generated_artifacts[0].name.endswith(".xlsx")
            assert Path(controller.state.generated_artifacts[0].location).exists()

            # Now verify finish completes the workflow
            finish_event = controller.execute(AgentDecision(action="finish", done=True))
            assert finish_event.kind == ExecutionEventKind.TASK_COMPLETED
            assert controller.state.final_status == FinalStatus.COMPLETED


# ---------------------------------------------------------------------------
# 4. Chain-of-Thought Non-Exposure Invariant
# ---------------------------------------------------------------------------


class TestChainOfThoughtNonExposureInvariant:
    """Verify generated workbook, metadata, and outputs contain no chain-of-thought."""

    def test_no_chain_of_thought_in_artifact_or_output(self):
        with TemporaryDirectory() as tmpdir:
            cap = GenerateExcelCapability(output_dir=tmpdir)
            req = CapabilityRequest(
                capability_name="generate_excel",
                inputs={
                    "requested_calculation": "Average thickness calculation",
                    "source_data_reference": "inspection_readings.xlsx",
                    "result": [{"equipment_id": "EQ-001", "average": 4.5}],
                    "methodology": "Arithmetic mean of thickness values.",
                    "verification_status": "PASSED",
                },
            )

            result = cap.invoke(req)

            # 1. Output dict must not contain chain-of-thought
            output_json = json.dumps(result.output)
            assert "thought" not in output_json.lower()
            assert "reasoning" not in output_json.lower()
            assert "chain_of_thought" not in output_json.lower()

            # 2. Observation must not contain chain-of-thought
            obs_json = result.observations[0].model_dump_json()
            assert "chain_of_thought" not in obs_json.lower()

            # 3. Excel text content must not contain chain-of-thought
            wb = openpyxl.load_workbook(result.output["file_path"], data_only=True)
            for sheetname in wb.sheetnames:
                ws = wb[sheetname]
                for row in ws.iter_rows(values_only=True):
                    for cell in row:
                        if cell is not None:
                            text = str(cell).lower()
                            assert "chain of thought" not in text
                            assert "thinking:" not in text
                            assert "<thought>" not in text
