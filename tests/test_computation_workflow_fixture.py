"""Deterministic end-to-end synthetic fixture for Workflow B.

This is the executable Phase 6.7 demonstration of:

user request -> Agent intent -> spreadsheet inspection -> computation formulation
-> Coding Model -> sandbox -> observation -> correction -> verification -> deliverable.

It uses only a temporary synthetic workbook and deterministic in-memory model and
sandbox doubles.  The production capability boundaries remain unchanged: the
Controller owns state and every capability is invoked through the Broker.
"""

from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path

import openpyxl
import pytest

from aegis.agent import (
    AttachmentDescriptor,
    IntentAnalysisRequest,
    PlanGenerationRequest,
    RouterAgentRuntime,
    SandboxObservationLoop,
)
from aegis.broker import RegistryCapabilityBroker
from aegis.capabilities import (
    CapabilityKind,
    CapabilityRegistry,
    GenerateCodeCapability,
    GenerateExcelCapability,
    InspectSpreadsheetCapability,
    MockSandboxRunner,
    RunCodeCapability,
    SandboxResult,
    VerifyResultCapability,
)
from aegis.capabilities.base import Capability, CapabilityMetadata
from aegis.config import load_config
from aegis.orchestration import ExecutionController, ExecutionEventKind, WorkflowName
from aegis.router import MockModelProvider, ModelRegistry, ModelRouter
from aegis.schemas import AgentDecision, CapabilityRequest, CapabilityResult, CapabilityResultStatus, FinalStatus, TaskState, VerificationStatus
from aegis.skills import ComputationContext, build_code_generation_prompt, prepare_generate_code_inputs, prepare_run_code_inputs


@dataclass(frozen=True)
class SyntheticComputationFixture:
    """One stable source workbook and its independently known expected result."""

    user_goal: str
    workbook_path: Path
    expected_records: list[dict[str, object]]


@pytest.fixture
def synthetic_computation_fixture(tmp_path: Path) -> SyntheticComputationFixture:
    """Create one deterministic industrial readings workbook for the full workflow."""
    workbook_path = tmp_path / "synthetic_equipment_readings.xlsx"
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Inspection Readings"
    sheet.append(
        [
            "Equipment_ID",
            "Measured_Thickness",
            "Min_Acceptable_Thickness",
        ]
    )
    sheet.append(["EQ-001", 4.2, 4.0])
    sheet.append(["EQ-001", 4.6, 4.0])
    sheet.append(["EQ-002", 2.6, 3.0])
    sheet.append(["EQ-002", 2.8, 3.0])
    workbook.save(workbook_path)

    return SyntheticComputationFixture(
        user_goal=(
            "Calculate the average measured thickness for each equipment item "
            "and identify equipment below its minimum acceptable thickness."
        ),
        workbook_path=workbook_path,
        expected_records=[
            {
                "equipment_id": "EQ-001",
                "average_measured_thickness": 4.4,
                "min_acceptable_thickness": 4.0,
                "below_min_acceptable_thickness": False,
            },
            {
                "equipment_id": "EQ-002",
                "average_measured_thickness": 2.7,
                "min_acceptable_thickness": 3.0,
                "below_min_acceptable_thickness": True,
            },
        ],
    )


class _FinishCapability(Capability):
    """Minimal configured control capability used solely to complete the fixture."""

    @property
    def metadata(self) -> CapabilityMetadata:
        return CapabilityMetadata(
            name="finish",
            kind=CapabilityKind.CONTROL,
            description="Complete the deterministic synthetic computation fixture.",
            input_modalities=("spreadsheet",),
        )

    def execute(self, request: CapabilityRequest) -> CapabilityResult:
        return CapabilityResult(
            request_id=request.request_id,
            status=CapabilityResultStatus.SUCCEEDED,
        )


def test_synthetic_computation_fixture_runs_workflow_b_end_to_end(
    synthetic_computation_fixture: SyntheticComputationFixture,
    tmp_path: Path,
) -> None:
    """Run the complete bounded computation flow, including one correction cycle."""
    fixture = synthetic_computation_fixture
    config = load_config()
    router = ModelRouter(ModelRegistry(config.models))

    agent_call = 0

    def agent_response(_request) -> str:
        nonlocal agent_call
        agent_call += 1
        responses = [
            {
                "intent": "computation",
                "modality": "spreadsheet",
                "workflow": "computation",
                "summary": "The request requires a spreadsheet computation workflow.",
            },
            {
                "intent": "computation",
                "modality": "spreadsheet",
                "workflow": "computation",
                "summary": "Inspect readings, calculate averages, verify them, and deliver a workbook.",
                "steps": [
                    {"capability_name": "inspect_spreadsheet", "purpose": "Inspect the source workbook.", "inputs": {}},
                    {"capability_name": "generate_code", "purpose": "Formulate the requested averages and threshold comparison.", "inputs": {}},
                    {"capability_name": "run_code", "purpose": "Execute the calculation in the sandbox.", "inputs": {}},
                    {"capability_name": "verify_result", "purpose": "Apply deterministic checks.", "inputs": {}},
                    {"capability_name": "generate_excel", "purpose": "Create the verified calculation deliverable.", "inputs": {}},
                    {"capability_name": "finish", "purpose": "Complete the governed workflow.", "inputs": {}},
                ],
            },
            {
                "directive": "retry_correct",
                "summary": "The failed sandbox execution requires corrected code generation.",
                "proposed_action": {
                    "action": "generate_code",
                    "inputs": {},
                    "done": False,
                    "summary": "Regenerate the calculation using the sandbox error.",
                },
            },
            {
                "directive": "continue",
                "summary": "The corrected calculation is ready for sandbox execution.",
                "proposed_action": {
                    "action": "run_code",
                    "inputs": {},
                    "done": False,
                    "summary": "Execute the corrected calculation in the sandbox.",
                },
            },
            {
                "directive": "verify",
                "summary": "The successful computation should be verified deterministically.",
                "proposed_action": {
                    "action": "verify_result",
                    "inputs": {},
                    "done": False,
                    "summary": "Verify the structured computation output.",
                },
            },
            {
                "directive": "continue",
                "summary": "Verification passed; create the spreadsheet deliverable.",
                "proposed_action": {
                    "action": "generate_excel",
                    "inputs": {},
                    "done": False,
                    "summary": "Create the verified spreadsheet deliverable.",
                },
            },
            {
                "directive": "finish",
                "summary": "The verified deliverable is ready to complete.",
                "proposed_action": {
                    "action": "finish",
                    "inputs": {},
                    "done": True,
                    "summary": "Complete the workflow.",
                },
            },
        ]
        return json.dumps(responses[agent_call - 1])

    def coding_response(request) -> str:
        if "Correction Context" not in request.prompt:
            return 'raise KeyError("Min_Acceptable_Thickness")'
        return (
            "# CORRECTED_COMPUTATION\n"
            "import json\n"
            f"records = {fixture.expected_records!r}\n"
            "print(json.dumps(records))\n"
        )

    sandbox = MockSandboxRunner(
        result_factory=lambda code, _path: (
            SandboxResult(
                stdout=json.dumps(fixture.expected_records),
                exit_code=0,
            )
            if "CORRECTED_COMPUTATION" in code
            else SandboxResult(
                stderr="KeyError: 'Min_Acceptable_Thickness'",
                exit_code=1,
            )
        )
    )
    agent_provider = MockModelProvider(response_factory=agent_response)
    coding_provider = MockModelProvider(response_factory=coding_response)

    registry = CapabilityRegistry(config.capabilities)
    registry.register(InspectSpreadsheetCapability())
    registry.register(GenerateCodeCapability(router=router, providers={"local_ollama": coding_provider}))
    registry.register(RunCodeCapability(sandbox=sandbox))
    registry.register(VerifyResultCapability())
    registry.register(GenerateExcelCapability(output_dir=tmp_path / "deliverables"))
    registry.register(_FinishCapability())

    state = TaskState(
        user_goal=fixture.user_goal,
        attachments=[str(fixture.workbook_path)],
        max_retries=2,
        max_iterations=8,
    )
    controller = ExecutionController(
        state,
        WorkflowName.COMPUTATION,
        RegistryCapabilityBroker(registry),
    )
    agent = RouterAgentRuntime(config.agent, router, {"local_ollama": agent_provider})
    loop = SandboxObservationLoop(agent, controller)

    # User request -> Agent intent -> Controller-compatible plan.
    intent = agent.decide_intent(
        IntentAnalysisRequest(
            user_goal=fixture.user_goal,
            attachments=[
                AttachmentDescriptor(
                    name=fixture.workbook_path.name,
                    media_type=(
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    ),
                )
            ],
        )
    )
    plan = agent.generate_plan(
        PlanGenerationRequest(
            user_goal=fixture.user_goal,
            intent=intent.intent,
            modality=intent.modality,
            available_capabilities=[
                "inspect_spreadsheet",
                "generate_code",
                "run_code",
                "verify_result",
                "generate_excel",
                "finish",
            ],
        )
    )
    assert intent.workflow == WorkflowName.COMPUTATION
    assert [step.capability_name for step in plan.steps] == [
        "inspect_spreadsheet",
        "generate_code",
        "run_code",
        "verify_result",
        "generate_excel",
        "finish",
    ]

    # Spreadsheet inspection -> deterministic computation formulation -> Coding Model.
    controller.execute(
        AgentDecision(
            action="inspect_spreadsheet",
            inputs={"workbook": str(fixture.workbook_path)},
        )
    )
    inspection = controller.last_capability_result
    assert inspection is not None
    assert inspection.output["columns"] == [
        "Equipment_ID",
        "Measured_Thickness",
        "Min_Acceptable_Thickness",
    ]

    context = ComputationContext(
        user_goal=fixture.user_goal,
        file_path=str(fixture.workbook_path),
        sheet_names=inspection.output["sheet_names"],
        columns=inspection.output["columns"],
        numeric_fields=inspection.output["numeric_fields"],
        row_count=inspection.output["row_count"],
        representative_values=inspection.output["representative_values"],
    )
    formulation = build_code_generation_prompt(context)
    assert "Measured_Thickness" in formulation.data_schema
    assert "Min_Acceptable_Thickness" in formulation.data_schema

    controller.execute(
        AgentDecision(
            action="generate_code",
            inputs=prepare_generate_code_inputs(formulation),
        )
    )
    first_generation = controller.last_capability_result
    assert first_generation is not None
    first_code = first_generation.output["code"]
    assert isinstance(first_code, str)
    assert "KeyError" in first_code
    assert first_generation.observations[0].data["model_id"] == "coding_model"

    # Sandbox -> observation -> Agent-directed correction -> corrected sandbox execution.
    failed_run = controller.execute(
        AgentDecision(
            action="run_code",
            inputs=prepare_run_code_inputs(first_code, str(fixture.workbook_path)),
        )
    )
    assert failed_run.kind == ExecutionEventKind.ACTION_FAILED
    assert controller.observation_for_agent().source == "run_code"

    recovery = loop.recover_from_run_code_failure(context, first_code)
    assert recovery.rerun_event is not None
    assert recovery.rerun_event.kind == ExecutionEventKind.ACTION_COMPLETED
    successful_run = controller.last_capability_result
    assert successful_run is not None
    assert successful_run.output["stdout"] == json.dumps(fixture.expected_records)

    # Agent-directed deterministic verification.
    verification = loop.reason()
    verification_event = loop.apply(
        verification,
        inputs_overlay={
            "stdout": successful_run.output["stdout"],
            "stderr": successful_run.output["stderr"],
            "exit_code": successful_run.output["exit_code"],
            "timed_out": successful_run.output["timed_out"],
            "computation_objective": fixture.user_goal,
            "expected_fields": [
                "equipment_id",
                "average_measured_thickness",
                "below_min_acceptable_thickness",
            ],
            "min_row_count": 2,
        },
    )
    assert verification_event.kind == ExecutionEventKind.ACTION_COMPLETED
    verification_result = controller.last_capability_result
    assert verification_result is not None
    assert verification_result.output["verified"] is True
    assert controller.state.verification_status == VerificationStatus.PASSED

    # Agent-directed local deliverable generation -> finish.
    deliver = loop.reason()
    deliver_event = loop.apply(
        deliver,
        inputs_overlay={
            "requested_calculation": fixture.user_goal,
            "source_data_reference": str(fixture.workbook_path),
            "stdout": successful_run.output["stdout"],
            "methodology": (
                "Arithmetic mean of Measured_Thickness grouped by Equipment_ID, "
                "compared with Min_Acceptable_Thickness."
            ),
            "verification_status": verification_result.output["summary"],
        },
    )
    assert deliver_event.kind == ExecutionEventKind.ACTION_COMPLETED
    assert len(controller.state.generated_artifacts) == 1
    deliverable = Path(controller.state.generated_artifacts[0].location)
    assert deliverable.exists()

    finish = loop.reason()
    finish_event = loop.apply(finish)
    assert finish_event.kind == ExecutionEventKind.TASK_COMPLETED

    # Verifiable complete-flow evidence without exposing model chain-of-thought.
    assert controller.state.final_status == FinalStatus.COMPLETED
    assert controller.state.retry_count == 1
    assert controller.state.iteration_count == 8
    assert sandbox.call_count == 2
    assert len(coding_provider.requests) == 2
    assert "Correction Context" in coding_provider.requests[1].prompt
    assert [
        event.action
        for event in controller.execution_events
        if event.kind == ExecutionEventKind.ACTION_STARTED
    ] == [
        "inspect_spreadsheet",
        "generate_code",
        "run_code",
        "generate_code",
        "run_code",
        "verify_result",
        "generate_excel",
        "finish",
    ]

    workbook = openpyxl.load_workbook(deliverable, data_only=True)
    assert workbook.sheetnames == ["Calculation Summary", "Detailed Results"]
    assert workbook["Calculation Summary"]["B7"].value == verification_result.output["summary"]
    assert workbook["Detailed Results"]["A3"].value == "EQ-002"
